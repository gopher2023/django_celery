from django.shortcuts import render
from django.shortcuts import HttpResponse
# Create your views here.
from books.models import Book
import json
import xlrd
from mycelery.books.tasks import get_books_data, app
from celery.result import AsyncResult
from openpyxl import Workbook

from django.utils.timezone import now
import datetime
def index_page(request):
    # ret =  Book.objects.all()
    # data = []
    # for item in ret:
    #     data.append({"title":item.title, "date":str(item.publication_date)})
    # data = json.dumps(data, ensure_ascii=False)

    return render(request, 'index_page.html')
def book_celery(request):
    task = get_books_data.delay()
    result = AsyncResult(id=task.id, app=app)
    print(result.status)
    print("----->", result.get())
    data = result.get()

    from io import BytesIO
    wb = Workbook()
    wb.encoding = 'utf-8'
    sheet1 = wb.active  # 获取第一个工作表（sheet1）
    sheet1.title = 'baobiao'  # 给工作表1设置标题

    sheet1.cell(row=1, column=1).value = 'books_nums'
    sheet1.cell(row=1, column=2).value = 'author_num'
    sheet1.cell(row=2, column=1).value = data.get("books_nums", 0)
    sheet1.cell(row=2, column=2).value = data.get("authors_num", 0)

    sheet1.cell(row=3, column=1).value = "detail info"
    sheet1.cell(row=4, column=1).value = 'author'
    sheet1.cell(row=4, column=2).value = 'num'
    row_num = 5
    for k, v in data["author_detail"].items():
        print(k, v)
        sheet1.cell(row=row_num, column=1).value = k
        sheet1.cell(row=row_num, column=2).value = v
        row_num += 1

    # # 准备写入到IO中
    output = BytesIO()
    wb.save(output)  # 将Excel文件内容保存到IO中
    output.seek(0)  # 重新定位到开始
    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
    ctime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    file_name = f'baobiao{ctime}.xls' # 给文件名中添加日期时间
    # file_name = urlquote(file_name)  # 使用urlquote()方法解决中文无法使用的问题
    response['Content-Disposition'] = 'attachment; filename=%s' % file_name
    response.write(output.getvalue())	 # 在设置HttpResponse的类型时，如果给了值，可以不写这句
    return response

def xl_import(request):
    if request.method == "POST":
        file_obj = request.FILES.get("fileUpload")
        type_excel = file_obj.name.split('.')[1]
        if 'xls' == type_excel:
            try:
                wb = xlrd.open_workbook(filename=None, file_contents=file_obj.read())
                table = wb.sheets()[0]
                nrows = table.nrows
                for i in range(1, nrows):
                    title = table.row_values(i)[0]
                    author = table.row_values(i)[1]
                    publication_date = xlrd.xldate_as_datetime(table.row_values(i)[2], 0).strftime('%Y-%m-%d')
                    print(publication_date)
                    if not Book.objects.filter(title=title).exists():
                        Book(title=title, author=author, publication_date=publication_date).save()
                    else:
                        Book.objects.filter(title=title).update(author=author, publication_date=publication_date)
                return HttpResponse("Success: 批量更新完成")
            except Exception as e:
                return HttpResponse(f"Error:{e}")

        return HttpResponse("hello world")
    if request.method == "GET":
        return render(request, "xl_import.html")